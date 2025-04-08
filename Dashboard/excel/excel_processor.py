import pandas as pd
from openpyxl import load_workbook

def process_excel_document(file_path, input_sheet_name="Q61a_DCN_J4_J46_Body_PartDoc"):
    """
    Reads an Excel sheet, consolidates text by document_id, and writes results to a new sheet.
    
    Parameters:
    file_path (str): Path to the Excel file
    input_sheet_name (str): Name of the sheet containing the raw data
    """
    try:
        # Read the Excel file into a pandas DataFrame
        # We specifically select columns A and C using usecols parameter
        df = pd.read_excel(file_path, sheet_name=input_sheet_name, usecols=[0, 2])  # A and C columns
        
        # Rename columns to ensure we're working with the correct names
        df.columns = ['document_id', 'document_text']
        
        # Group by document_id and aggregate the text
        # We use lambda function to join texts with spaces while preserving order
        processed_df = df.groupby('document_id', as_index=False).agg({
            'document_text': lambda x: ' '.join(str(text).strip() for text in x)
        })
        
        # Load the existing workbook to add the new sheet
        book = load_workbook(file_path)
        
        # Create a new sheet name
        new_sheet_name = "Processed_Data"
        
        # If the sheet already exists, remove it
        if new_sheet_name in book.sheetnames:
            book.remove(book[new_sheet_name])
        
        # Save and close the workbook to prevent file lock
        book.save(file_path)
        book.close()
        
        # Write the processed DataFrame to a new sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            processed_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
        
        print(f"Processing complete! Results written to sheet '{new_sheet_name}'")
        print(f"Processed {len(df)} rows into {len(processed_df)} consolidated documents")
        
        # Return both DataFrames for inspection if needed
        return df, processed_df
    
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None, None

def preview_data(original_df, processed_df):
    """
    Prints a preview of both original and processed data for verification.
    """
    print("\nOriginal Data Preview:")
    print("----------------------")
    print(original_df.head())
    
    print("\nProcessed Data Preview:")
    print("----------------------")
    print(processed_df.head())
    
    # Print some statistics
    print("\nStatistics:")
    print(f"Number of original rows: {len(original_df)}")
    print(f"Number of unique document IDs: {len(processed_df)}")
    
# Example usage
if __name__ == "__main__":
    # Replace with your Excel file path
    file_path = r'C:\Users\A488466\Desktop\Q61a-DCN-Body-InPipe_June24.xlsx'
    
    # Process the document
    original_df, processed_df = process_excel_document(file_path)
    
    # If processing was successful, show the preview
    if original_df is not None and processed_df is not None:
        preview_data(original_df, processed_df)